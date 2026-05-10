"""Custom tools for the Excel Research Agent.

Provides web search (DuckDuckGo, no API key needed) and Excel generation.
"""

from __future__ import annotations

import os
from typing import Any

from langchain_core.tools import tool
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _ensure_str(value: Any) -> str:
    """Coerce a value to a string safely."""
    if value is None:
        return ""
    return str(value)


@tool
def web_search(query: str, max_results: int = 5) -> list[dict[str, str]]:
    """Search the web using DuckDuckGo and return results.

    Args:
        query: Search query string.
        max_results: Maximum number of results to return (default 5).

    Returns:
        List of search results with title, url, and snippet.
    """
    try:
        from ddgs import DDGS
    except ImportError as exc:
        msg = "ddgs is not installed. Run: uv add ddgs"
        raise ImportError(msg) from exc

    with DDGS() as ddgs:
        results = list(ddgs.text(query, max_results=max_results))

    return [
        {
            "title": _ensure_str(r.get("title")),
            "url": _ensure_str(r.get("href")),
            "snippet": _ensure_str(r.get("body")),
        }
        for r in results
    ]


@tool
def generate_excel(
    data: list[dict[str, Any]],
    filename: str,
    sheet_name: str = "Data",
) -> str:
    """Generate an Excel file from structured data.

    Args:
        data: List of dictionaries where each dict is a row.
            Keys become column headers.
        filename: Output filename (e.g., 'report.xlsx').
            Saved to the current working directory.
        sheet_name: Name of the worksheet (default 'Data').

    Returns:
        Absolute path to the generated Excel file.
    """
    if not data:
        return "Error: No data provided"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(data[0].keys())
    header_fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid",
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(_ensure_str(cell.value)))
            except Exception:  # noqa: S110  # best-effort width calc
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    output_path = os.path.abspath(filename)
    wb.save(output_path)
    return f"Excel file saved to: {output_path}"
